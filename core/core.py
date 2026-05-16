# app.py
# FastAPI 服务：上传 xlsx，生成 30个/页、居中、等宽对折、四线三格 PDF

import io
import math
import re
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.utils import ImageReader

app = FastAPI(title="英语默写 PDF 生成服务")

# 这个字体是 reportlab 内置 CID 字体，不需要额外分发字体文件
pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

FONT_CJK = "STSong-Light"
FONT_EN = "Helvetica"
BLACK = colors.black

# 建议把你那个“黑色虚线.pdf”里截出来的一行四线三格图，保存成 static/fourline.png
# 也可以在初始化时从 PDF 提取，但实际部署更推荐固定成图片资源。
GRID_IMAGE_PATH = Path("static/fourline.png")


def clean_text(value):
    text = str(value or "")
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def read_words_from_xlsx(file_bytes: bytes, sheet_name: str | None = None):
    """
    默认按之前文件格式读取：
    A列：序号
    B列：英文
    D列：中文释义
    前两行是标题，从第3行开始读取。
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            idx = int(row[0])
        except Exception:
            continue
        english = clean_text(row[1] if len(row) > 1 else "")
        chinese = clean_text(row[3] if len(row) > 3 else "")
        rows.append((idx, english, chinese))
    return rows


def string_width(text, font, size):
    return pdfmetrics.stringWidth(str(text), font, size)


def tokenize(text):
    text = str(text or "")
    if not text:
        return []

    # 中文释义：中文按字切，英文/数字/符号按块切，方便自动换行
    if re.search(r"[\u4e00-\u9fff]", text):
        pattern = re.compile(r"[A-Za-z0-9./&()\-]+|[\u4e00-\u9fff]|[^A-Za-z0-9\u4e00-\u9fff]")
        raw = pattern.findall(text)
        tokens = []
        prev_space = False
        for t in raw:
            if t.isspace():
                if not prev_space:
                    tokens.append(" ")
                prev_space = True
            else:
                tokens.append(t)
                prev_space = False
        return tokens

    # 英文短语：优先按空格换行
    parts = text.split(" ")
    tokens = []
    for i, p in enumerate(parts):
        if i:
            tokens.append(" ")
        tokens.append(p)
    return tokens


def wrap_text(text, width, font, size, max_lines=2):
    text = str(text or "")
    if not text:
        return []

    lines = []
    current = ""
    for token in tokenize(text):
        candidate = current + token
        if string_width(candidate, font, size) <= width or not current:
            current = candidate
        else:
            lines.append(current.strip())
            current = token.strip() if token != " " else ""
            if len(lines) >= max_lines:
                break

    if current and len(lines) < max_lines:
        lines.append(current.strip())
    return lines[:max_lines]


def fit_lines(text, width, font, start_size, min_size=5.8, max_lines=3):
    """字号从 start_size 逐步缩小，直到不超宽；仍超出就按 max_lines 换行。"""
    size = start_size
    while size >= min_size:
        lines = wrap_text(text, width, font, size, max_lines=max_lines)
        if not lines or max(string_width(line, font, size) for line in lines) <= width:
            return lines, size
        size -= 0.3
    return wrap_text(text, width, font, min_size, max_lines=max_lines), min_size


def draw_centered_text(c, text, x, y, w, h, font, size, max_lines=2, min_size=6.0):
    padding = 1.0 * mm
    lines, actual_size = fit_lines(text, w - 2 * padding, font, size, min_size, max_lines)
    if not lines:
        return

    leading = actual_size * 1.15
    total_h = leading * len(lines)
    start_y = y + (h + total_h) / 2 - actual_size

    c.setFont(font, actual_size)
    for i, line in enumerate(lines):
        line_w = string_width(line, font, actual_size)
        c.drawString(x + (w - line_w) / 2, start_y - i * leading, line)


def draw_fourline_grid(c, x, y, w, h, grid_reader):
    """
    直接贴原版四线三格图片，保证和参考 PDF 的线条深浅、虚线、斜线一致。
    图片建议由“黑色虚线.pdf”裁出一行后保存为 static/fourline.png。
    """
    draw_w = w - 2 * mm
    draw_h = 6.0 * mm   # 四线总高度：1.5 + 3 + 1.5 = 6mm
    draw_x = x + 1.0 * mm
    draw_y = y + (h - draw_h) / 2
    c.drawImage(grid_reader, draw_x, draw_y, width=draw_w, height=draw_h, preserveAspectRatio=False, mask="auto")


def build_pdf(rows, rows_per_page=30):
    """
    生成 PDF 并返回 bytes。
    版式：A4、30个一页、字体居中、长文本换行、序号窄列，其余四列等宽，方便对折。
    """
    if not GRID_IMAGE_PATH.exists():
        raise FileNotFoundError("缺少 static/fourline.png，请先从黑色虚线.pdf 裁出四线三格样式图。")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4, pageCompression=1)

    page_w, page_h = A4
    table_w = 190 * mm
    idx_w = 8 * mm
    other_w = (table_w - idx_w) / 4
    col_widths = [idx_w, other_w, other_w, other_w, other_w]
    headers = ["序号", "英文", "默写汉语", "中文", "默写英文"]

    left = (page_w - table_w) / 2
    top_margin = 10 * mm
    header_h = 7 * mm
    row_h = 8.5 * mm
    grid_reader = ImageReader(str(GRID_IMAGE_PATH))

    total_pages = math.ceil(len(rows) / rows_per_page)

    for page_index in range(total_pages):
        page_rows = rows[page_index * rows_per_page:(page_index + 1) * rows_per_page]

        table_x = left
        table_top = page_h - top_margin
        y_header = table_top - header_h
        table_h = header_h + row_h * rows_per_page
        table_bottom = table_top - table_h

        xs = [table_x]
        for cw in col_widths:
            xs.append(xs[-1] + cw)

        # 表格线
        c.setStrokeColor(BLACK)
        c.setLineWidth(0.5)
        for xx in xs:
            c.line(xx, table_bottom, xx, table_top)
        c.line(table_x, table_top, table_x + table_w, table_top)
        c.line(table_x, y_header, table_x + table_w, y_header)
        for i in range(rows_per_page + 1):
            yy = y_header - i * row_h
            c.line(table_x, yy, table_x + table_w, yy)

        # 表头
        c.setFont(FONT_CJK, 9.2)
        for i, header in enumerate(headers):
            tw = string_width(header, FONT_CJK, 9.2)
            c.drawString(xs[i] + (col_widths[i] - tw) / 2, y_header + (header_h - 9.2) / 2 + 1.2, header)

        # 默写英文列：每行贴原版四线三格
        for r_i in range(rows_per_page):
            y_top = y_header - r_i * row_h
            y = y_top - row_h
            draw_fourline_grid(c, xs[4], y, col_widths[4], row_h, grid_reader)

        # 内容
        for r_i, (idx, english, chinese) in enumerate(page_rows):
            y_top = y_header - r_i * row_h
            y = y_top - row_h

            draw_centered_text(c, str(idx), xs[0], y, col_widths[0], row_h, FONT_EN, 8.3, max_lines=1, min_size=6.4)
            draw_centered_text(c, english, xs[1], y, col_widths[1], row_h, FONT_EN, 9.3, max_lines=2, min_size=6.4)
            draw_centered_text(c, chinese, xs[3], y, col_widths[3], row_h, FONT_CJK, 8.8, max_lines=3, min_size=5.7)

        # 页码
        c.setFont(FONT_EN, 7.5)
        c.drawCentredString(page_w / 2, 7.2 * mm, str(page_index + 1))
        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer.getvalue()


@app.post("/generate")
async def generate_pdf(
    file: UploadFile = File(...),
    rows_per_page: int = Form(30),
    sheet_name: str | None = Form(None),
):
    file_bytes = await file.read()
    rows = read_words_from_xlsx(file_bytes, sheet_name=sheet_name)
    pdf_bytes = build_pdf(rows, rows_per_page=rows_per_page)

    output_name = "英语单词默写练习_30个一页_等宽对折版.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{output_name}"},
    )


@app.get("/")
def index():
    return {"message": "POST /generate 上传 xlsx 生成 PDF"}
